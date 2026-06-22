# app/documents/parsers/xlsx_parser.py
#
# XLSX extraction via openpyxl. One Document per sheet, rendered as
# a markdown table with the first row as header. read_only=True
# since we only ever read uploaded files, never write them back —
# significantly reduces memory usage on large spreadsheets.
#
# Known limitation: merged cells read as None/empty for every cell
# except the top-left anchor — openpyxl does not auto-fill merged
# values. Acceptable for v1; revisit if real documents make this a
# practical problem.

import io
import logging

from langchain_core.documents import Document
from openpyxl import load_workbook

logger = logging.getLogger("app.documents.parsers.xlsx")


def parse_xlsx(file_bytes: bytes, filename: str) -> list[Document]:
    """One Document per sheet."""
    workbook = load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True
    )
    documents = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = [
            [str(cell) if cell is not None else "" for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]

        if not rows:
            logger.info(
                "%s sheet '%s' is empty — skipped", filename, sheet_name
            )
            continue

        header = "| " + " | ".join(rows[0]) + " |"
        separator = "| " + " | ".join("---" for _ in rows[0]) + " |"
        body = "\n".join("| " + " | ".join(row) + " |" for row in rows[1:])
        markdown_table = "\n".join([header, separator, body])

        documents.append(
            Document(
                page_content=markdown_table,
                metadata={
                    "source": filename,
                    "file_type": "xlsx",
                    "sheet": sheet_name,
                },
            )
        )

    workbook.close()
    return documents